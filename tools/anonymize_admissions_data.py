#!/usr/bin/env python
"""
Local-only anonymizer for admissions campaign exports.

The script reads CSV/XLSX files from disk and writes an anonymized copy.
It does not use network access, external APIs, telemetry, or cloud services.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - optional dependency for CSV-only runs
    openpyxl = None


PERSON_NAME_RE = re.compile(r"фио|фамил|имя|отчест|full.?name|name", re.I)
PHONE_RE = re.compile(r"тел|phone|mobile|мобил|номер", re.I)
EMAIL_RE = re.compile(r"почт|email|e-mail|mail", re.I)
PASSPORT_RE = re.compile(r"паспорт|снилс|snils|инн|inn|документ|серия|номер.?док", re.I)
BIRTH_RE = re.compile(r"рожд|birth", re.I)
ADDRESS_RE = re.compile(r"адрес|улица|дом|кварт|address|прожив|регистрац", re.I)
COMMENT_RE = re.compile(r"коммент|примеч|note|remark|описан|сообщен|текст", re.I)
ID_RE = re.compile(r"(^id$|guid|uid|uuid|код|идентификатор|уник)", re.I)
DATE_RE = re.compile(r"дата|date", re.I)


@dataclass
class ColumnRule:
    column: str
    action: str


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_hash(value: str, salt: str, length: int = 12) -> str:
    payload = (salt + "|" + value).encode("utf-8", errors="ignore")
    return hashlib.sha256(payload).hexdigest()[:length]


def classify_column(column: str) -> str:
    if PERSON_NAME_RE.search(column):
        return "person_name"
    if PHONE_RE.search(column):
        return "phone"
    if EMAIL_RE.search(column):
        return "email"
    if PASSPORT_RE.search(column):
        return "sensitive_document"
    if BIRTH_RE.search(column):
        return "birth_date"
    if ADDRESS_RE.search(column):
        return "address"
    if COMMENT_RE.search(column):
        return "free_text"
    if ID_RE.search(column):
        return "identifier"
    if DATE_RE.search(column):
        return "event_date"
    return "keep"


def parse_possible_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = norm(value)
    if not text:
        return None
    formats = [
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%y",
        "%d/%m/%y",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def row_key(row: dict[str, Any], row_number: int) -> str:
    preferred = []
    for key, value in row.items():
        if classify_column(key) in {"identifier", "person_name", "email", "phone"} and norm(value):
            preferred.append(norm(value))
    if preferred:
        return "|".join(preferred)
    return f"row:{row_number}"


def anon_id_for(row: dict[str, Any], row_number: int, salt: str) -> str:
    digest = stable_hash(row_key(row, row_number), salt, length=10).upper()
    return f"ABIT-{digest[:4]}-{digest[4:]}"


def mask_phone(value: Any, anon_id: str) -> str:
    if not norm(value):
        return ""
    suffix = stable_hash(anon_id, "phone", length=4)
    return f"+7-XXX-XXX-{suffix}"


def mask_email(value: Any, anon_id: str) -> str:
    if not norm(value):
        return ""
    return f"{anon_id.lower().replace('-', '_')}@example.local"


def shift_event_date(value: Any, anon_id: str) -> Any:
    dt = parse_possible_date(value)
    if not dt:
        return value
    shift = (int(stable_hash(anon_id, "date-shift", length=2), 16) % 15) - 7
    return (dt + timedelta(days=shift)).strftime("%Y-%m-%d")


def birth_year(value: Any) -> str:
    dt = parse_possible_date(value)
    if dt:
        return str(dt.year)
    text = norm(value)
    match = re.search(r"(19|20)\d{2}", text)
    return match.group(0) if match else ("[год скрыт]" if text else "")


def anonymize_value(value: Any, rule: str, anon_id: str, applicant_alias: str) -> Any:
    text = norm(value)
    if not text:
        return ""
    if rule == "person_name":
        return applicant_alias
    if rule == "phone":
        return mask_phone(value, anon_id)
    if rule == "email":
        return mask_email(value, anon_id)
    if rule == "sensitive_document":
        return "[удалено]"
    if rule == "birth_date":
        return birth_year(value)
    if rule == "address":
        return "[адрес обезличен]"
    if rule == "free_text":
        return "[текст обезличен]"
    if rule == "identifier":
        return anon_id
    if rule == "event_date":
        return shift_event_date(value, anon_id)
    return value


def detect_csv_dialect(path: Path) -> csv.Dialect:
    sample = path.read_text(encoding=detect_encoding(path), errors="replace")[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,	")
    except csv.Error:
        class Fallback(csv.excel):
            delimiter = ";"
        return Fallback


def detect_encoding(path: Path) -> str:
    raw = path.read_bytes()[:4]
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            path.read_text(encoding=enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "utf-8"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]], csv.Dialect, str]:
    encoding = detect_encoding(path)
    dialect = detect_csv_dialect(path)
    with path.open("r", encoding=encoding, newline="", errors="replace") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames or []), rows, dialect, encoding


def write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]], dialect: csv.Dialect) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter=getattr(dialect, "delimiter", ";"))
        writer.writeheader()
        writer.writerows(rows)


def read_xlsx(path: Path, sheet_name: str | None) -> tuple[list[str], list[dict[str, Any]], str]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for XLSX files. Use CSV or install openpyxl locally.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [norm(cell) for cell in next(rows_iter)]
    except StopIteration:
        return [], [], ws.title
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        rows.append({header[i]: values[i] if i < len(values) else "" for i in range(len(header))})
    return header, rows, ws.title


def write_xlsx(path: Path, columns: list[str], rows: list[dict[str, Any]], sheet_name: str) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for XLSX files. Use CSV or install openpyxl locally.")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "anonymized")[:31]
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    wb.save(path)


def anonymize_rows(columns: list[str], rows: list[dict[str, Any]], salt: str) -> tuple[list[str], list[dict[str, Any]], list[ColumnRule]]:
    rules = [ColumnRule(column=column, action=classify_column(column)) for column in columns]
    output_columns = ["anon_applicant_id"] + columns
    output_rows: list[dict[str, Any]] = []
    applicant_aliases: dict[str, str] = {}
    for idx, row in enumerate(rows, 1):
        anon_id = anon_id_for(row, idx, salt)
        if anon_id not in applicant_aliases:
            applicant_aliases[anon_id] = f"Абитуриент {len(applicant_aliases) + 1:04d}"
        applicant_alias = applicant_aliases[anon_id]
        next_row: dict[str, Any] = {"anon_applicant_id": anon_id}
        for rule in rules:
            next_row[rule.column] = anonymize_value(row.get(rule.column), rule.action, anon_id, applicant_alias)
        output_rows.append(next_row)
    return output_columns, output_rows, rules


def write_report(path: Path, source: Path, output: Path, row_count: int, rules: Iterable[ColumnRule]) -> None:
    report = {
        "source_file": str(source),
        "output_file": str(output),
        "rows": row_count,
        "note": "Report contains schema and transformations only; no source values are included.",
        "columns": [{"name": rule.column, "action": rule.action} for rule in rules],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(input_path.stem + "_anonymized" + input_path.suffix)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Anonymize admissions campaign CSV/XLSX locally.")
    parser.add_argument("input", help="Path to source CSV/XLSX file with personal data.")
    parser.add_argument("-o", "--output", help="Path to anonymized output file.")
    parser.add_argument("--sheet", help="XLSX sheet name. Defaults to first sheet.")
    parser.add_argument("--salt", default=os.environ.get("ANONYMIZER_SALT", "pk360-local-only"), help="Stable local salt. Change it if you need a different pseudonym mapping.")
    parser.add_argument("--report", help="Path to JSON report with transformation rules.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else default_output_path(input_path)
    suffix = input_path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        columns, rows, dialect, _encoding = read_csv(input_path)
        output_columns, output_rows, rules = anonymize_rows(columns, rows, args.salt)
        write_csv(output_path, output_columns, output_rows, dialect)
    elif suffix in {".xlsx", ".xlsm"}:
        columns, rows, sheet_name = read_xlsx(input_path, args.sheet)
        output_columns, output_rows, rules = anonymize_rows(columns, rows, args.salt)
        write_xlsx(output_path, output_columns, output_rows, sheet_name)
    else:
        raise SystemExit(f"Unsupported file type: {suffix}. Use CSV, TSV, XLSX or XLSM.")

    report_path = Path(args.report).expanduser().resolve() if args.report else output_path.with_suffix(output_path.suffix + ".report.json")
    write_report(report_path, input_path, output_path, len(output_rows), rules)
    print(f"Anonymized rows: {len(output_rows)}")
    print(f"Output: {output_path}")
    print(f"Report: {report_path}")
    print("Local-only run complete. No network calls were made by this script.")


if __name__ == "__main__":
    main()
