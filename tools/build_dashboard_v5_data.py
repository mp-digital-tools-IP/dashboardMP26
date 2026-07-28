"""Build privacy-safe aggregates for Dashboard V5.

The source workbook is already anonymized, but this script still never exports
names, phones, e-mails, personal file numbers or free-text fields.  The browser
receives only aggregates and a small deterministic anonymous sample.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


FORM_LABELS = {
    "Очная": "Очная",
    "Очно-заочная": "Очно-заочная",
    "Заочная": "Заочная",
}
FUNDING_LABELS = {
    "Федеральный бюджет": "budget",
    "Внебюджетные средства": "paid",
}
LEVEL_LABELS = {
    "Бакалавриат": "Бакалавриат и специалитет",
    "Специалитет": "Бакалавриат и специалитет",
    "Магистратура": "Магистратура",
    "Аспирантура": "Аспирантура",
}
OFFICIAL_TOTALS = {
    "Бакалавриат и специалитет": {
        "Все формы": {"budget": 2854, "paid": 5369},
        "Очная": {"budget": 2373, "paid": 3665},
        "Очно-заочная": {"budget": 151, "paid": 690},
        "Заочная": {"budget": 330, "paid": 1014},
    },
    "Магистратура": {"Все формы": {"budget": 779, "paid": 1953}},
    "Аспирантура": {"Все формы": {"budget": 66, "paid": 0}},
}

EXCLUDED_FACULTY_TOKENS = ("филиал",)


def text(value) -> str:
    return "" if value is None else str(value).strip()


def number(value) -> int | None:
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return None


def priority(value) -> int:
    result = number(value)
    return result if result is not None else 999


def normalized(value: str) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", " ", value.casefold()).strip()


def person_key(case_value, fallback) -> str:
    match = re.search(r"Личное дело\s+(\d+)", text(case_value))
    return match.group(1) if match else text(fallback)


def anonymous_id(value: str) -> str:
    digest = hashlib.sha256(f"pk26:{value}".encode("utf-8")).hexdigest().upper()
    return f"AB-26-{digest[:6]}"


def application_date(case_value) -> date | None:
    match = re.search(r"от\s+(\d{2}\.\d{2}\.\d{4})", text(case_value))
    return datetime.strptime(match.group(1), "%d.%m.%Y").date() if match else None


def parse_plan_cell(value: str) -> dict[str, int]:
    value = value.replace("—", "").strip()
    if not value:
        return {"budget": 0, "paid": 0}
    parts = [part.strip() for part in value.split("/")]
    if len(parts) == 1:
        return {"budget": number(parts[0]) or 0, "paid": 0}
    return {"budget": number(parts[0]) or 0, "paid": number(parts[1]) or 0}


def parse_official_programs(path: Path) -> list[dict]:
    programs: list[dict] = []
    current_level = ""
    code_re = re.compile(r"^\d{2}\.\d{2}\.\d{2}(?:\.\d{2})?$")
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if line == "### Специалитет":
            current_level = "Специалитет"
            continue
        if line == "### Бакалавриат":
            current_level = "Бакалавриат"
            continue
        if line.startswith("### ") and line not in {"### Специалитет", "### Бакалавриат"}:
            current_level = ""
        if not current_level or not line.startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) < 5 or not code_re.match(cells[0]):
            continue
        code, name = cells[0], cells[1]
        programs.append({
            "code": code,
            "groupCode": ".".join(code.split(".")[:3]),
            "name": name,
            "level": current_level,
            "plans": {
                "Очная": parse_plan_cell(cells[2]),
                "Очно-заочная": parse_plan_cell(cells[3]),
                "Заочная": parse_plan_cell(cells[4]),
            },
            "search": normalized(name),
        })
    return programs


def best_program(profile: str, direction: str, level: str, official: list[dict], fallback: dict[str, str]):
    profile_n = normalized(profile)
    candidates = [item for item in official if item["level"] == level]
    exact = [item for item in candidates if profile_n and (profile_n in item["search"] or item["search"] in profile_n)]
    if exact:
        return max(exact, key=lambda item: min(len(profile_n), len(item["search"])))
    direction_code = fallback.get(direction)
    if direction_code:
        return next((item for item in candidates if item["code"] == direction_code), {
            "code": direction_code,
            "groupCode": ".".join(direction_code.split(".")[:3]),
            "name": profile or direction,
            "level": level,
            "plans": {form: {"budget": 0, "paid": 0} for form in FORM_LABELS},
        })
    return {
        "code": "—",
        "groupCode": "—",
        "name": profile or direction or "Не указано",
        "level": level,
        "plans": {form: {"budget": 0, "paid": 0} for form in FORM_LABELS},
    }


def new_stats():
    return {
        "rows": 0,
        "people": set(),
        "applications": set(),
        "activePeople": set(),
        "consentPeople": set(),
        "activeConsentPeople": set(),
        "potentialBudgetPeople": set(),
        "potentialPaidPeople": set(),
        "budgetRows": 0,
        "paidRows": 0,
        "scores": [],
    }


def add_stats(stats, person, app_key, active, consent, basis, score, priority_value):
    stats["rows"] += 1
    stats["people"].add(person)
    stats["applications"].add(app_key)
    if active:
        stats["activePeople"].add(person)
    if consent:
        stats["consentPeople"].add(person)
        if active:
            stats["activeConsentPeople"].add(person)
    if basis == "budget" and active and priority_value <= 2:
        stats["potentialBudgetPeople"].add(person)
    if basis == "paid" and active and priority_value <= 2:
        stats["potentialPaidPeople"].add(person)
    if basis == "budget":
        stats["budgetRows"] += 1
    elif basis == "paid":
        stats["paidRows"] += 1
    if score is not None:
        stats["scores"].append(score)


def finish_stats(stats, high_ids):
    values = sorted(stats["scores"])
    median = values[len(values) // 2] if values else None
    return {
        "rows": stats["rows"],
        "people": len(stats["people"]),
        "applications": len(stats["applications"]),
        "activePeople": len(stats["activePeople"]),
        "consentPeople": len(stats["consentPeople"]),
        "activeConsentPeople": len(stats["activeConsentPeople"]),
        "potentialBudgetPeople": len(stats["potentialBudgetPeople"]),
        "potentialPaidPeople": len(stats["potentialPaidPeople"]),
        "budgetRows": stats["budgetRows"],
        "paidRows": stats["paidRows"],
        "medianScore": median,
        "highScorers": len(stats["people"].intersection(high_ids)),
        "highScorerConsents": len(stats["consentPeople"].intersection(high_ids)),
        "activeHighScorerConsents": len(stats["activeConsentPeople"].intersection(high_ids)),
    }


def cumulative_series(person_dates, application_dates, paid_dates):
    available = list(person_dates.values()) + list(application_dates.values()) + list(paid_dates.values())
    if not available:
        return []
    first, last = min(available), max(available)
    people_daily = Counter(person_dates.values())
    apps_daily = Counter(application_dates.values())
    paid_daily = Counter(paid_dates.values())
    current = first
    people = applications = paid = 0
    output = []
    while current <= last:
        people += people_daily[current]
        applications += apps_daily[current]
        paid += paid_daily[current]
        output.append({"date": current.strftime("%d.%m.%Y"), "people": people, "applications": applications, "paid": paid})
        current += timedelta(days=1)
    return output


def modeled_touchpoints(person: str, apps: list[dict], has_consent: bool) -> list[dict]:
    """Create a deterministic, explicitly modeled CRM/event journey for the demo."""
    digest = hashlib.sha256(f"touches:{person}".encode("utf-8")).digest()
    app_dates = []
    for item in apps:
        if item.get("date"):
            try:
                app_dates.append(datetime.strptime(item["date"], "%d.%m.%Y").date())
            except ValueError:
                pass
    anchor = min(app_dates) if app_dates else date(2026, 7, 1)
    count = (3 if has_consent else 1) + digest[0] % 3
    event_names = [
        "День открытых дверей",
        "Профориентационное мероприятие",
        "Выставка «Образование и карьера»",
        "Консультация по поступлению",
    ]
    call_results = [
        "Дозвонились · интерес подтверждён",
        "Обсудили приоритеты и конкурс",
        "Запрошена консультация по программе",
        "Направлена памятка о следующем шаге",
    ]
    touches = []
    for index in range(count):
        offset = 54 - index * 11 - digest[index + 1] % 7
        touch_date = anchor - timedelta(days=max(2, offset))
        if index % 2 == 0:
            event_name = event_names[digest[index + 4] % len(event_names)]
            touches.append({
                "type": "Участие в мероприятии",
                "source": "Модель CRM",
                "date": touch_date.strftime("%d.%m.%Y"),
                "result": event_name,
                "modeled": True,
            })
        else:
            touches.append({
                "type": "Звонок",
                "source": "Модель CRM",
                "date": touch_date.strftime("%d.%m.%Y"),
                "result": call_results[digest[index + 4] % len(call_results)],
                "modeled": True,
            })
    return sorted(touches, key=lambda item: datetime.strptime(item["date"], "%d.%m.%Y"))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("knowledge", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--v4-actuals", type=Path)
    args = parser.parse_args()

    official = parse_official_programs(args.knowledge)
    fallback = {}
    if args.v4_actuals and args.v4_actuals.exists():
        old = json.loads(args.v4_actuals.read_text(encoding="utf-8"))
        fallback = {item["name"]: item["code"] for item in old.get("directions", []) if item.get("code") != "—"}

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator)
    col = {name: index for index, name in enumerate(headers)}
    required = [
        "anon_applicant_id", "Личное дело", "Сумма баллов", "Согласие на зачисление", "Приоритет",
        "Форма обучения", "Источник финансирования", "Уровень подготовки", "Направление\\специальность",
        "Профиль", "Подразделение", "Текущий статус конкурса",
    ]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"Missing columns: {missing}")

    all_people = set()
    subject_scores = defaultdict(dict)
    scope_stats = defaultdict(new_stats)
    faculty_stats = defaultdict(new_stats)
    program_stats = defaultdict(new_stats)
    program_meta = {}
    ranking = defaultdict(dict)
    applicant_apps = defaultdict(list)
    applicant_meta = {}
    programs_by_person = defaultdict(set)
    dynamics = defaultdict(lambda: {"people": {}, "applications": {}, "paid": {}})
    match_cache = {}

    for row in iterator:
        raw_id = text(row[col["anon_applicant_id"]])
        if not raw_id:
            continue
        person = person_key(row[col["Личное дело"]], raw_id)
        level_raw = text(row[col["Уровень подготовки"]])
        level_group = LEVEL_LABELS.get(level_raw)
        if not level_group:
            continue
        form = FORM_LABELS.get(text(row[col["Форма обучения"]]), text(row[col["Форма обучения"]]) or "Не указано")
        basis = FUNDING_LABELS.get(text(row[col["Источник финансирования"]]), "other")
        direction = text(row[col["Направление\\специальность"]])
        profile = text(row[col["Профиль"]])
        faculty = text(row[col["Подразделение"]]) or "Не указано"
        if any(token in faculty.casefold() for token in EXCLUDED_FACULTY_TOKENS):
            continue
        all_people.add(person)
        cache_key = (profile, direction, level_raw)
        if cache_key not in match_cache:
            match_cache[cache_key] = best_program(profile, direction, level_raw, official, fallback)
        program = match_cache[cache_key]
        code = program["code"]
        program_id = f"{level_group}|{code}|{program['name']}"
        program_meta[program_id] = {**program, "id": program_id, "levelGroup": level_group, "faculty": faculty}
        active = text(row[col["Текущий статус конкурса"]]) == "Участвует в конкурсе"
        consent = text(row[col["Согласие на зачисление"]]).casefold() == "да"
        priority_value = priority(row[col["Приоритет"]])
        score = number(row[col["Сумма баллов"]])
        row_date = application_date(row[col["Личное дело"]])
        app_key = (person, program_id)

        for form_key in {form, "Все формы"}:
            for basis_key in {basis, "all"}:
                add_stats(scope_stats[(level_group, form_key, basis_key)], person, app_key, active, consent, basis, score, priority_value)
                add_stats(faculty_stats[(level_group, form_key, basis_key, faculty)], person, app_key, active, consent, basis, score, priority_value)
                add_stats(program_stats[(program_id, form_key, basis_key)], person, app_key, active, consent, basis, score, priority_value)

        if row_date:
            for form_key in {form, "Все формы"}:
                for dynamic_key in [
                    f"scope:{level_group}:{form_key}",
                    f"faculty:{level_group}:{form_key}:{faculty}",
                    f"program:{form_key}:{program_id}",
                ]:
                    bucket = dynamics[dynamic_key]
                    old_person = bucket["people"].get(person)
                    if old_person is None or row_date < old_person:
                        bucket["people"][person] = row_date
                    old_app = bucket["applications"].get(app_key)
                    if old_app is None or row_date < old_app:
                        bucket["applications"][app_key] = row_date
                    if basis == "paid":
                        old_paid = bucket["paid"].get(app_key)
                        if old_paid is None or row_date < old_paid:
                            bucket["paid"][app_key] = row_date

        for index in range(1, 11):
            discipline = text(row[col.get(f"Дисциплина{index}", -1)]) if f"Дисциплина{index}" in col else ""
            subject = number(row[col.get(f"Предмет{index}", -1)]) if f"Предмет{index}" in col else None
            if discipline and subject is not None:
                subject_scores[person][discipline] = max(subject, subject_scores[person].get(discipline, 0))

        rank_key = (person, form, basis)
        existing = ranking[program_id].get(rank_key)
        candidate = {
            "id": anonymous_id(person), "score": score or 0, "priority": priority_value,
            "consent": consent, "active": active, "status": text(row[col["Текущий статус конкурса"]]) or "Не указан",
            "form": form, "basis": basis,
        }
        if existing is None or (candidate["priority"], -candidate["score"]) < (existing["priority"], -existing["score"]):
            ranking[program_id][rank_key] = candidate
        programs_by_person[person].add(program_id)
        existing_app = next((
            item for item in applicant_apps[person]
            if item["programId"] == program_id and item["form"] == form and item["basis"] == basis
        ), None)
        if existing_app is None and len(applicant_apps[person]) < 12:
            applicant_apps[person].append({
                "programId": program_id, "code": code, "name": program["name"], "faculty": faculty,
                "score": score or 0, "priority": candidate["priority"], "consent": consent, "active": active,
                "status": candidate["status"], "form": form, "basis": basis,
                "date": row_date.strftime("%d.%m.%Y") if row_date else None,
            })
        elif existing_app is not None:
            existing_app["score"] = max(existing_app["score"], score or 0)
            existing_app["priority"] = min(existing_app["priority"], candidate["priority"])
            existing_app["consent"] = existing_app["consent"] or consent
            was_active = existing_app["active"]
            existing_app["active"] = existing_app["active"] or active
            if consent or (active and not was_active):
                existing_app["status"] = candidate["status"]
            if not existing_app["date"] and row_date:
                existing_app["date"] = row_date.strftime("%d.%m.%Y")
        meta = applicant_meta.setdefault(person, {"id": anonymous_id(person), "maxScore": 0, "faculties": set(), "active": False, "consent": False, "priorityReady": False})
        meta["maxScore"] = max(meta["maxScore"], score or 0)
        meta["faculties"].add(faculty)
        meta["active"] = meta["active"] or active
        meta["consent"] = meta["consent"] or consent
        meta["priorityReady"] = meta["priorityReady"] or (basis == "budget" and active and priority_value <= 2)

    def subject_value(values, tokens):
        matches = [score for name, score in values.items() if any(token in name.casefold() for token in tokens)]
        return max(matches) if matches else None

    high_ids = set()
    for person, values in subject_scores.items():
        math = subject_value(values, ["математ"])
        physics = subject_value(values, ["физик"])
        informatics = subject_value(values, ["информат"])
        if math is not None and math >= 85 and ((physics is not None and physics >= 85) or (informatics is not None and informatics >= 85)):
            high_ids.add(person)

    best_budget_consent = {}
    for program_id, candidates in ranking.items():
        for (person, _form, basis), item in candidates.items():
            if basis != "budget" or not item["active"] or not item["consent"]:
                continue
            choice = (item["priority"], -item["score"], program_id)
            if person not in best_budget_consent or choice < best_budget_consent[person]:
                best_budget_consent[person] = choice
    budget_filled_by_program = Counter(choice[2] for choice in best_budget_consent.values())

    scope_output = {}
    for key, stats in scope_stats.items():
        level_group, form, basis = key
        result = finish_stats(stats, high_ids)
        total = OFFICIAL_TOTALS.get(level_group, {}).get(form) or OFFICIAL_TOTALS.get(level_group, {}).get("Все формы", {"budget": 0, "paid": 0})
        result["planBudget"] = total["budget"] if basis in {"all", "budget"} else 0
        result["planPaid"] = total["paid"] if basis in {"all", "paid"} else 0
        scope_output["|".join(key)] = result

    faculties_output = []
    faculty_pairs = sorted({(key[0], key[3]) for key in faculty_stats})
    for level_group, faculty in faculty_pairs:
        base = finish_stats(faculty_stats[(level_group, "Все формы", "all", faculty)], high_ids)
        slices = {}
        for form in ["Все формы", *FORM_LABELS]:
            for basis in ["all", "budget", "paid"]:
                stats = faculty_stats.get((level_group, form, basis, faculty))
                if stats:
                    slices[f"{form}|{basis}"] = finish_stats(stats, high_ids)
        base.update({"name": faculty, "level": level_group, "slices": slices})
        faculties_output.append(base)
    faculties_output.sort(key=lambda item: (item["level"], -item["applications"]))

    programs_output = []
    ranking_output = {}
    for program_id, meta in program_meta.items():
        stats = program_stats[(program_id, "Все формы", "all")]
        result = finish_stats(stats, high_ids)
        plan_budget = sum(item["budget"] for item in meta["plans"].values())
        plan_paid = sum(item["paid"] for item in meta["plans"].values())
        potential_candidates = [
            item for item in ranking[program_id].values()
            if item["basis"] == "budget" and item["active"] and item["priority"] <= 2 and item["score"] > 0
        ]
        potential_candidates.sort(key=lambda item: (-item["score"], item["priority"], item["id"]))
        projected_top = potential_candidates[:plan_budget] if plan_budget else []
        top_ids = {item["id"] for item in projected_top}
        result.update({
            "id": program_id, "code": meta["code"], "groupCode": meta["groupCode"], "name": meta["name"],
            "faculty": meta["faculty"], "level": meta["levelGroup"], "plans": meta["plans"],
            "planBudget": plan_budget, "planPaid": plan_paid,
            "projectedTopCount": min(plan_budget, len(potential_candidates)),
            "projectedTopRate": round(min(plan_budget, len(potential_candidates)) / plan_budget * 100, 1) if plan_budget else None,
            "projectedAverageScore": round(sum(item["score"] for item in projected_top) / len(projected_top), 1) if projected_top else None,
            "projectedBoundaryScore": projected_top[-1]["score"] if projected_top else None,
            "budgetFilled": min(plan_budget, budget_filled_by_program[program_id]) if plan_budget else 0,
            "budgetFillRate": round(min(plan_budget, budget_filled_by_program[program_id]) / plan_budget * 100, 1) if plan_budget else None,
            "topAverageScore": round(sum(item["score"] for item in projected_top if item["consent"]) / len([item for item in projected_top if item["consent"]]), 1) if any(item["consent"] for item in projected_top) else None,
            "topBoundaryScore": next((item["score"] for item in reversed(projected_top) if item["consent"]), None),
            "slices": {
                f"{form}|{basis}": finish_stats(program_stats[(program_id, form, basis)], high_ids)
                for form in ["Все формы", *FORM_LABELS]
                for basis in ["all", "budget", "paid"]
                if (program_id, form, basis) in program_stats
            },
        })
        programs_output.append(result)
        rows = sorted(ranking[program_id].values(), key=lambda item: (-item["score"], item["priority"], item["id"]))[:80]
        for place, item in enumerate(rows, 1):
            item["place"] = place
            item["topList"] = item["id"] in top_ids
        ranking_output[program_id] = rows
    programs_output.sort(key=lambda item: (-item["applications"], item["code"]))

    intersections = Counter()
    for person_programs in programs_by_person.values():
        values = sorted(person_programs)
        if len(values) > 12:
            values = values[:12]
        for index, left in enumerate(values):
            for right in values[index + 1:]:
                intersections[(left, right)] += 1
    intersection_output = [{"a": a, "b": b, "count": count} for (a, b), count in intersections.most_common(80)]

    ranked_applicants = sorted(
        applicant_meta.items(),
        key=lambda item: (-int(item[1]["consent"]), -int(item[0] in high_ids), -int(item[1]["priorityReady"]), -item[1]["maxScore"], item[1]["id"]),
    )
    def balanced_sample(candidates, limit):
        by_faculty = defaultdict(list)
        for person, meta in candidates:
            for faculty in sorted(meta["faculties"]):
                by_faculty[faculty].append((person, meta))
        selected = []
        selected_people = set()
        faculty_names = sorted(by_faculty)
        while len(selected) < limit:
            added = False
            for faculty in faculty_names:
                while by_faculty[faculty] and by_faculty[faculty][0][0] in selected_people:
                    by_faculty[faculty].pop(0)
                if by_faculty[faculty]:
                    person, meta = by_faculty[faculty].pop(0)
                    selected_people.add(person)
                    selected.append((person, meta))
                    added = True
                    if len(selected) >= limit:
                        break
            if not added:
                break
        if len(selected) < limit:
            selected.extend((person, meta) for person, meta in candidates if person not in selected_people)
        return selected[:limit]

    with_consent = [item for item in ranked_applicants if item[1]["consent"]]
    without_consent = [item for item in ranked_applicants if not item[1]["consent"]]
    applicant_candidates = balanced_sample(with_consent, 120) + balanced_sample(without_consent, 120)
    applicants_output = []
    for person, meta in applicant_candidates:
        apps = sorted(applicant_apps[person], key=lambda item: (item["priority"], -item["score"]))
        segment = "Высокобалльник 85+" if person in high_ids else "Приоритет 1–2" if meta["priorityReady"] else "Активен в конкурсе" if meta["active"] else "Требует внимания"
        exams = [{"name": name, "score": score} for name, score in sorted(subject_scores[person].items(), key=lambda item: (-item[1], item[0]))]
        raw_consent_apps = [item for item in apps if item["consent"]]
        consent_app = min(
            raw_consent_apps,
            key=lambda item: (not item["active"], item["priority"], item["basis"] != "budget", -item["score"], item["programId"]),
            default=None,
        )
        for item in apps:
            item["consent"] = item is consent_app
        consent_programs = [{
            "programId": consent_app["programId"], "code": consent_app["code"], "name": consent_app["name"],
            "faculty": consent_app["faculty"], "form": consent_app["form"], "basis": consent_app["basis"],
            "active": consent_app["active"],
        }] if consent_app else []
        touchpoints = modeled_touchpoints(person, apps, bool(consent_programs))
        applicants_output.append({
            "id": meta["id"], "score": meta["maxScore"], "faculties": sorted(meta["faculties"]),
            "segment": segment, "applications": apps, "consent": bool(consent_programs),
            "consentPrograms": consent_programs,
            "exams": exams, "examTotal": sum(item["score"] for item in exams),
            "touchpoints": touchpoints, "touchCount": len(touchpoints),
            "touchStatus": "Модель демонстрационной траектории; не связана с фактической CRM",
        })

    dynamics_output = {
        key: cumulative_series(bucket["people"], bucket["applications"], bucket["paid"])
        for key, bucket in dynamics.items()
    }

    result = {
        "source": {"file": args.input.name, "rows": sheet.max_row - 1, "updated": "27.07.2026", "branchesExcluded": True, "excludedScope": "Все институты и филиалы вне московской площадки"},
        "officialTotals": OFFICIAL_TOTALS,
        "allExport": {"rows": sheet.max_row - 1, "people": len(all_people)},
        "definitions": {
            "people": "Уникальные номера личных дел",
            "applications": "Уникальная пара «человек × образовательная программа»",
            "highScorer": "Математика 85+ и физика либо информатика 85+",
            "potentialBudget": "Люди с активным бюджетным заявлением и приоритетом 1–2; это сигнал намерения, а не согласие на зачисление",
            "topList": "Расчётный верхний диапазон по активным бюджетным заявлениям с приоритетом 1–2; согласие показывается отдельным фактическим признаком",
            "model2025": "Визуализационная модель: основной поток 90→94% от 2026, платный поток 88→93%",
        },
        "scopes": scope_output,
        "faculties": faculties_output,
        "programs": programs_output,
        "dynamics": dynamics_output,
        "rankings": ranking_output,
        "applicants": applicants_output,
        "intersections": intersection_output,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "rows": result["allExport"]["rows"], "people": result["allExport"]["people"],
        "programs": len(programs_output), "faculties": len({item["name"] for item in faculties_output}),
        "rankingGroups": len(ranking_output), "applicantSample": len(applicants_output),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
